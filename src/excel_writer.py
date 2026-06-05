"""Build the daily Excel tracker of jobs to apply to."""
from __future__ import annotations
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER = [
    "Priority", "Match %", "Company", "Role", "Location", "Remote",
    "Salary", "Source", "Posted", "Apply Link",
    "Resume File", "Cover Letter File", "Status", "Notes", "Why This Job",
]


def build_tracker(jobs: list[dict], out_path: Path, day: str):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Apps {day}"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, name in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data
    for i, j in enumerate(jobs, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=j.get("match_score", ""))
        ws.cell(row=row, column=3, value=j["company"])
        ws.cell(row=row, column=4, value=j["title"])
        ws.cell(row=row, column=5, value=j.get("location", ""))
        ws.cell(row=row, column=6, value="Yes" if j.get("remote") else "No")
        ws.cell(row=row, column=7, value=j.get("salary", ""))
        ws.cell(row=row, column=8, value=j.get("source", ""))
        posted = j.get("posted_at", "")
        if isinstance(posted, datetime):
            posted = posted.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row, column=9, value=posted)

        link_cell = ws.cell(row=row, column=10, value="Apply")
        if j.get("url"):
            link_cell.hyperlink = j["url"]
            link_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=row, column=11, value=j.get("resume_file", ""))
        ws.cell(row=row, column=12, value=j.get("cover_file", ""))
        status_cell = ws.cell(row=row, column=13, value="To Apply")
        status_cell.font = Font(color="B8860B", bold=True)
        ws.cell(row=row, column=14, value="")
        ws.cell(row=row, column=15, value=j.get("match_reason", ""))

        for col in range(1, len(HEADER) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical="center", wrap_text=(col in (4, 5, 15))
            )

    # Column widths tuned for readability
    widths = [7, 8, 22, 32, 22, 7, 14, 20, 16, 10, 30, 34, 10, 28, 48]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Status validation dropdown
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"To Apply,Applied,Interview,Offer,Rejected,Skip"',
        allow_blank=True,
    )
    dv.add(f"M2:M{len(jobs) + 1}")
    ws.add_data_validation(dv)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
